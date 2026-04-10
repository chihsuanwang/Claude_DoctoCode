"""
create_excel_examples.py
========================
產生 Equity Futures Options 的 Excel 計算範例（共 4 個檔案）：
  01_European_Futures_Option.xlsx   — Black-76 解析解
  02_American_Futures_Option_BAW.xlsx — Barone-Adesi Whaley
  03_Quanto_Futures_Option.xlsx     — Quanto（固定匯率）
  04_CrossCurrency_Futures_Option.xlsx — 跨幣種選擇權

執行：
  python create_excel_examples.py
"""

import os
import numpy as np
from scipy.stats import norm
from scipy.optimize import brentq

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series

# ─────────────────────────────────────────────────────────────
# 輸出目錄
# ─────────────────────────────────────────────────────────────
OUTPUT_DIR = r"c:\Users\chihs\Claude_DoctoCode\文件"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────────────────────
# 色彩 & 樣式常數
# ─────────────────────────────────────────────────────────────
C_DARK   = "1F497D"   # 深藍（標題）
C_WHITE  = "FFFFFF"
C_LBLUE  = "D6E4F0"   # 淡藍（區塊標題）
C_INPUT  = "FFF2CC"   # 淡黃（輸入格）
C_CALC   = "F5F5F5"   # 淡灰（中間計算）
C_OUT    = "E2EFDA"   # 淡綠（輸出）
C_BORD   = "B0B0B0"   # 灰色框線

FN = "微軟正黑體"

def _b(color=C_BORD, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(color):    return PatternFill("solid", fgColor=color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name=FN, bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", indent=0, wrap=False):
    return Alignment(horizontal=h, vertical=v, indent=indent, wrap_text=wrap)

def apply(cell, fill=None, font=None, align=None, num_fmt=None, border=True):
    if fill:   cell.fill   = _fill(fill)
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    if border: cell.border = _b()

def set_val(ws, row, col, value, **kwargs):
    c = ws.cell(row=row, column=col, value=value)
    apply(c, **kwargs)
    return c

# ── 複合列書寫器 ────────────────────────────────────────────

def title_row(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    apply(c, fill=C_DARK, font=_font(bold=True, size=14, color=C_WHITE),
          align=_align("left", indent=1), border=False)
    ws.row_dimensions[row].height = 30

def subtitle_row(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    apply(c, fill="E8F0FB", font=_font(size=9, color="444444", italic=True),
          align=_align("left", indent=1), border=False)

def sec_header(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    apply(c, fill=C_LBLUE, font=_font(bold=True, size=10, color=C_DARK),
          align=_align("left", indent=1))
    ws.row_dimensions[row].height = 18

def inp_row(ws, row, label, symbol, value, unit="", note=""):
    """輸入列（淡黃底，粗體數值）"""
    set_val(ws, row, 1, label, font=_font(size=10), fill=C_INPUT)
    set_val(ws, row, 2, symbol, font=_font(bold=True, italic=True, color=C_DARK),
            fill=C_INPUT, align=_align("center"))
    c = set_val(ws, row, 3, value, font=_font(bold=True, size=10, color="C00000"),
                fill=C_INPUT, align=_align("right"))
    set_val(ws, row, 4, unit, font=_font(size=9, color="666666"), fill=C_INPUT)
    if note:
        nc = ws.cell(row=row, column=5, value=note)
        apply(nc, font=_font(size=9, color="888888", italic=True), border=False)
    return c

def calc_row(ws, row, label, symbol, formula, unit="", is_out=False):
    """計算列（灰底或綠底）"""
    bg = C_OUT if is_out else C_CALC
    set_val(ws, row, 1, label, font=_font(size=10, bold=is_out), fill=bg)
    set_val(ws, row, 2, symbol,
            font=_font(bold=True, italic=True, color=C_DARK, size=10),
            fill=bg, align=_align("center"))
    c = ws.cell(row=row, column=3, value=formula)
    apply(c, fill=bg, font=_font(size=10, bold=is_out, color=C_DARK if is_out else "333333"),
          align=_align("right"), num_fmt="#,##0.0000")
    set_val(ws, row, 4, unit, font=_font(size=9, color="666666"), fill=bg)
    return c  # return formula cell

def note_row(ws, row, text, cols=5):
    """備注列（無框線，斜體）"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    apply(c, font=_font(size=9, italic=True, color="888888"),
          align=_align("left", indent=1), border=False)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─────────────────────────────────────────────────────────────
# 金融模型函式（Python 計算，用於敏感度表格預算）
# ─────────────────────────────────────────────────────────────

def black76(F, K, r, T, sigma):
    """Black-76: European option on futures. Returns dict."""
    if T <= 0 or sigma <= 0:
        call = max(F - K, 0.0)
        put  = max(K - F, 0.0)
        return {"call": call, "put": put, "d1": 0, "d2": 0, "df": 1,
                "delta_c": 1 if F>K else 0, "delta_p": -1 if F<K else 0,
                "gamma": 0, "vega": 0, "theta_c": 0, "theta_p": 0}
    d1  = (np.log(F/K) + 0.5*sigma**2*T) / (sigma*np.sqrt(T))
    d2  = d1 - sigma*np.sqrt(T)
    df  = np.exp(-r*T)
    Nd1, Nd2   = norm.cdf(d1),  norm.cdf(d2)
    Nmd1, Nmd2 = norm.cdf(-d1), norm.cdf(-d2)
    phi_d1 = norm.pdf(d1)
    call = df*(F*Nd1  - K*Nd2)
    put  = df*(K*Nmd2 - F*Nmd1)
    delta_c =  df*Nd1
    delta_p = -df*Nmd1
    gamma   =  df*phi_d1 / (F*sigma*np.sqrt(T))
    vega    =  F*df*phi_d1*np.sqrt(T)*0.01   # per 1%
    theta_c = (-F*df*phi_d1*sigma/(2*np.sqrt(T)) - r*call)/365
    theta_p = (-F*df*phi_d1*sigma/(2*np.sqrt(T)) + r*put)/365
    return dict(call=call, put=put, d1=d1, d2=d2, df=df,
                delta_c=delta_c, delta_p=delta_p,
                gamma=gamma, vega=vega, theta_c=theta_c, theta_p=theta_p)


def _baw_gamma_c(r, T, sigma):
    """BAW γ_c parameter (for futures: n=0)"""
    m = 2*r*T / (sigma**2)
    ell = 1 - np.exp(-r*T)
    if ell < 1e-12:
        return 1e10
    return (1 + np.sqrt(1 + 4*m/ell)) / 2

def _baw_gamma_p(r, T, sigma):
    """BAW γ_p parameter"""
    m = 2*r*T / (sigma**2)
    ell = 1 - np.exp(-r*T)
    if ell < 1e-12:
        return -1e10
    return (1 - np.sqrt(1 + 4*m/ell)) / 2


def _find_fstar_call(K, r, T, sigma):
    """Find F*_c by bisection (Brent's method)"""
    if r < 1e-8:
        return np.inf
    gamma_c = _baw_gamma_c(r, T, sigma)

    def g(x):
        d1 = (np.log(x/K) + 0.5*sigma**2*T)/(sigma*np.sqrt(T))
        c  = np.exp(-r*T)*(x*norm.cdf(d1) - K*norm.cdf(d1 - sigma*np.sqrt(T)))
        Ac = (x/gamma_c)*(1 - np.exp(-r*T)*norm.cdf(d1))
        return x - K - c - Ac

    lo, hi = K + 1e-6, K * 500
    try:
        return brentq(g, lo, hi, xtol=1e-7)
    except Exception:
        return hi


def _find_fstar_put(K, r, T, sigma):
    """Find F*_p by bisection"""
    if r < 1e-8:
        return 0.0
    gamma_p = _baw_gamma_p(r, T, sigma)

    def g(x):
        d1 = (np.log(x/K) + 0.5*sigma**2*T)/(sigma*np.sqrt(T))
        p  = np.exp(-r*T)*(K*norm.cdf(-(d1 - sigma*np.sqrt(T))) - x*norm.cdf(-d1))
        Ap = (x/gamma_p)*(np.exp(-r*T)*norm.cdf(-d1) - 1)
        return K - x - p - Ap

    lo, hi = 1e-6, K - 1e-6
    try:
        return brentq(g, lo, hi, xtol=1e-7)
    except Exception:
        return lo


def baw(F, K, r, T, sigma):
    """Barone-Adesi Whaley: American option on futures."""
    eur = black76(F, K, r, T, sigma)
    gamma_c = _baw_gamma_c(r, T, sigma)
    gamma_p = _baw_gamma_p(r, T, sigma)
    fstar_c = _find_fstar_call(K, r, T, sigma)
    fstar_p = _find_fstar_put(K, r, T, sigma)

    # Call
    if F >= fstar_c:
        call_price = F - K
    else:
        d1_fstar = (np.log(fstar_c/K) + 0.5*sigma**2*T)/(sigma*np.sqrt(T))
        Ac = (fstar_c/gamma_c)*(1 - np.exp(-r*T)*norm.cdf(d1_fstar))
        call_price = eur["call"] + Ac*(F/fstar_c)**gamma_c

    # Put
    if F <= fstar_p:
        put_price = K - F
    else:
        d1_fstar_p = (np.log(fstar_p/K) + 0.5*sigma**2*T)/(sigma*np.sqrt(T))
        Ap = (fstar_p/gamma_p)*(np.exp(-r*T)*norm.cdf(-d1_fstar_p) - 1)
        put_price = eur["put"] + Ap*(F/fstar_p)**gamma_p

    return dict(call=call_price, put=put_price,
                fstar_c=fstar_c, fstar_p=fstar_p,
                gamma_c=gamma_c, gamma_p=gamma_p,
                eur_call=eur["call"], eur_put=eur["put"])


def quanto(F, K, r_d, T, sigma, sigma_x, rho, X_bar=1.0, option_type="call"):
    """
    Quanto option on futures (fixed FX rate).
    Drift adjustment: F_adj = F * exp(-rho*sigma*sigma_x*T)
    Price in payoff currency = (1/X_bar) * Black76(F_adj, K, r_d, T, sigma)
    """
    F_adj = F * np.exp(-rho * sigma * sigma_x * T)
    if T <= 0 or sigma <= 0:
        px = max(F_adj - K, 0) / X_bar if option_type == "call" else max(K - F_adj, 0) / X_bar
        return {"price": px, "F_adj": F_adj}
    d1 = (np.log(F_adj/K) + 0.5*sigma**2*T) / (sigma*np.sqrt(T))
    d2 = d1 - sigma*np.sqrt(T)
    df = np.exp(-r_d*T)
    if option_type == "call":
        px = df*(F_adj*norm.cdf(d1) - K*norm.cdf(d2)) / X_bar
    else:
        px = df*(K*norm.cdf(-d2) - F_adj*norm.cdf(-d1)) / X_bar
    return {"price": px, "F_adj": F_adj, "d1": d1, "d2": d2, "df": df}


def cross_currency(F, K, r_d, T, sigma, sigma_x, rho, X0, option_type="call"):
    """
    Cross-currency option on futures.
    Effective vol: sigma_eff = sqrt(sigma^2 + sigma_x^2 + 2*rho*sigma*sigma_x)
    Price in payoff ccy = e^{-r_d*T} * [(F/X0)*N(d1) - K*N(d2)]
    where d1 uses Z0=F/X0 and sigma_eff.
    """
    sigma_eff = np.sqrt(sigma**2 + sigma_x**2 + 2*rho*sigma*sigma_x)
    Z0 = F / X0   # futures price in payoff currency
    if T <= 0 or sigma_eff <= 0:
        px = max(Z0 - K, 0) if option_type == "call" else max(K - Z0, 0)
        return {"price": px, "sigma_eff": sigma_eff, "Z0": Z0}
    d1 = (np.log(Z0/K) + 0.5*sigma_eff**2*T) / (sigma_eff*np.sqrt(T))
    d2 = d1 - sigma_eff*np.sqrt(T)
    df = np.exp(-r_d*T)
    if option_type == "call":
        px = df*(Z0*norm.cdf(d1) - K*norm.cdf(d2))
    else:
        px = df*(K*norm.cdf(-d2) - Z0*norm.cdf(-d1))
    return {"price": px, "sigma_eff": sigma_eff, "Z0": Z0, "d1": d1, "d2": d2, "df": df}


# ─────────────────────────────────────────────────────────────
# Excel 產生：01 European (Black-76)
# ─────────────────────────────────────────────────────────────

def create_european(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Black-76 Calculator"
    set_col_widths(ws, [32, 8, 16, 12, 40])

    # ── 標題 ──
    title_row(ws, 1, "European Futures Option — Black-76 Model")
    subtitle_row(ws, 2, "Source: Black (1976), Journal of Financial Economics. "
                 "Formula: C = e^(−rT)[F·N(d₁) − K·N(d₂)], P = e^(−rT)[K·N(−d₂) − F·N(−d₁)]")

    # ── 輸入 ──
    sec_header(ws, 4, "▌ INPUT PARAMETERS  （黃色格可修改）")
    inp_row(ws, 5,  "Futures Price",          "F",  100.00, "USD",    "目前期貨價格")
    inp_row(ws, 6,  "Strike Price",           "K",  100.00, "USD",    "選擇權執行價格")
    inp_row(ws, 7,  "Risk-Free Rate",         "r",  0.05,   "decimal","折現用無風險利率（如 5% → 0.05）")
    inp_row(ws, 8,  "Time to Expiry",         "T",  0.50,   "Years",  "到期年數")
    inp_row(ws, 9,  "Volatility",             "σ",  0.20,   "decimal","年化波動度（如 20% → 0.20）")
    inp_row(ws, 10, "Notional (# contracts)", "N",  1,      "units",  "合約數量（乘數）")

    # 常用 cell 參考
    F, K, r, T, s, N = "C5","C6","C7","C8","C9","C10"

    # ── 中間計算 ──
    sec_header(ws, 12, "▌ INTERMEDIATE CALCULATIONS")
    calc_row(ws, 13, "d₁",             "d₁",
             f"=(LN({F}/{K})+0.5*{s}^2*{T})/({s}*SQRT({T}))",  "")
    calc_row(ws, 14, "d₂",             "d₂",
             f"=C13-{s}*SQRT({T})",                              "")
    calc_row(ws, 15, "Discount Factor", "e^(−rT)",
             f"=EXP(-{r}*{T})",                                  "")
    calc_row(ws, 16, "N(d₁)",          "N(d₁)",
             "=NORM.S.DIST(C13,TRUE)",                           "")
    calc_row(ws, 17, "N(d₂)",          "N(d₂)",
             "=NORM.S.DIST(C14,TRUE)",                           "")
    calc_row(ws, 18, "N(−d₁)",         "N(−d₁)",
             "=NORM.S.DIST(-C13,TRUE)",                          "")
    calc_row(ws, 19, "N(−d₂)",         "N(−d₂)",
             "=NORM.S.DIST(-C14,TRUE)",                          "")
    calc_row(ws, 20, "φ(d₁) — normal PDF", "φ(d₁)",
             "=NORM.S.DIST(C13,FALSE)",                          "")

    # ── 選擇權價格 ──
    sec_header(ws, 22, "▌ OPTION PRICES")
    calc_row(ws, 23, "Call Price  C = e^(−rT)[F·N(d₁)−K·N(d₂)]",  "C",
             f"={N}*C15*({F}*C16-{K}*C17)", "USD", is_out=True)
    calc_row(ws, 24, "Put Price   P = e^(−rT)[K·N(−d₂)−F·N(−d₁)]", "P",
             f"={N}*C15*({K}*C19-{F}*C18)", "USD", is_out=True)
    calc_row(ws, 25, "Put-Call Parity Check  [C−P = e^(−rT)(F−K)]", "✓",
             f"=C23/{N}-C24/{N}-C15*({F}-{K})", "",  is_out=False)
    note_row(ws, 26, "  ☑ 值應接近 0（數值誤差 < 1e-10）")

    # ── Greeks ──
    sec_header(ws, 28, "▌ GREEKS  （per contract, N=1）")
    calc_row(ws, 29, "Delta — Call  Δ_c = e^(−rT)·N(d₁)",     "Δ_c",
             "=C15*C16",                                         "")
    calc_row(ws, 30, "Delta — Put   Δ_p = −e^(−rT)·N(−d₁)",   "Δ_p",
             "=-C15*C18",                                        "")
    calc_row(ws, 31, "Gamma  Γ = e^(−rT)·φ(d₁)/(F·σ·√T)",     "Γ",
             f"=C15*C20/({F}*{s}*SQRT({T}))",                  "per USD")
    calc_row(ws, 32, "Vega  ν = F·e^(−rT)·φ(d₁)·√T  (per 1% σ)", "ν",
             f"={F}*C15*C20*SQRT({T})*0.01",                   "USD / 1%σ")
    calc_row(ws, 33, "Theta — Call  (per calendar day)",         "Θ_c",
             f"=(-{F}*C15*C20*{s}/(2*SQRT({T}))-{r}*C15*({F}*C16-{K}*C17))/365",
             "USD / day")
    calc_row(ws, 34, "Theta — Put   (per calendar day)",         "Θ_p",
             f"=(-{F}*C15*C20*{s}/(2*SQRT({T}))+{r}*C15*({K}*C19-{F}*C18))/365",
             "USD / day")
    calc_row(ws, 35, "Rho — Call   (per 1% rate change)",        "ρ_c",
             f"=-{T}*(C23/{N})*0.01",                           "USD / 1%r")
    calc_row(ws, 36, "Rho — Put    (per 1% rate change)",        "ρ_p",
             f"=-{T}*(C24/{N})*0.01",                           "USD / 1%r")

    # ── 敏感度表格（Python 預算） ──
    sec_header(ws, 38, "▌ SENSITIVITY TABLE — Futures Price vs Option Price & Delta")
    # 參數固定
    pF  = np.arange(70, 135, 5)
    r0, T0, s0 = 0.05, 0.50, 0.20
    K0  = 100.0

    hdr_vals = ["Futures Price F"] + [f"{f:.0f}" for f in pF]
    for j, v in enumerate(hdr_vals):
        c = ws.cell(row=39, column=j+1, value=v)
        apply(c, fill=C_DARK, font=_font(bold=True, color=C_WHITE, size=9),
              align=_align("center"))

    rows_data = {
        "Call Price":   [black76(f, K0, r0, T0, s0)["call"]    for f in pF],
        "Put Price":    [black76(f, K0, r0, T0, s0)["put"]     for f in pF],
        "Call Delta":   [black76(f, K0, r0, T0, s0)["delta_c"] for f in pF],
        "Put Delta":    [black76(f, K0, r0, T0, s0)["delta_p"] for f in pF],
        "Gamma":        [black76(f, K0, r0, T0, s0)["gamma"]   for f in pF],
        "Vega (1%σ)":  [black76(f, K0, r0, T0, s0)["vega"]    for f in pF],
    }
    for ri, (lbl, vals) in enumerate(rows_data.items(), start=40):
        c = ws.cell(row=ri, column=1, value=lbl)
        apply(c, fill=C_CALC, font=_font(size=9))
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=ri, column=j, value=round(v, 4))
            apply(c, fill="FFFFFF" if ri % 2 == 0 else C_CALC,
                  font=_font(size=9), num_fmt="#,##0.0000")

    note_row(ws, 47, f"  ※ 固定參數：K={K0}, r={r0}, T={T0} yr, σ={s0}  (與上方輸入格無連動)")

    # ── Vol 敏感度 ──
    sec_header(ws, 49, "▌ SENSITIVITY TABLE — Volatility vs Option Price  (F=100, K=100, r=5%, T=0.5yr)")
    vol_range = np.arange(0.05, 0.55, 0.05)
    h2 = ["Volatility σ"] + [f"{v*100:.0f}%" for v in vol_range]
    for j, v in enumerate(h2):
        c = ws.cell(row=50, column=j+1, value=v)
        apply(c, fill=C_DARK, font=_font(bold=True, color=C_WHITE, size=9), align=_align("center"))
    for ri, lbl in enumerate(["Call Price", "Put Price"], start=51):
        ws.cell(row=ri, column=1, value=lbl)
        apply(ws.cell(row=ri, column=1), fill=C_CALC, font=_font(size=9))
        for j, sv in enumerate(vol_range, start=2):
            res = black76(100, K0, r0, T0, sv)
            v = res["call"] if lbl == "Call Price" else res["put"]
            c = ws.cell(row=ri, column=j, value=round(v, 4))
            apply(c, fill="FFFFFF" if ri % 2 == 0 else C_CALC,
                  font=_font(size=9), num_fmt="#,##0.0000")

    # ── 圖表：Call & Put vs F ──
    ws2 = wb.create_sheet("Chart — Price vs F")
    # 複製價格資料到 Chart sheet
    ws2.cell(row=1, column=1, value="F")
    ws2.cell(row=1, column=2, value="Call")
    ws2.cell(row=1, column=3, value="Put")
    for i, f in enumerate(pF):
        res = black76(f, K0, r0, T0, s0)
        ws2.cell(row=i+2, column=1, value=float(f))
        ws2.cell(row=i+2, column=2, value=round(res["call"], 4))
        ws2.cell(row=i+2, column=3, value=round(res["put"],  4))

    chart = LineChart()
    chart.title = "European Futures Option Price vs Futures Price"
    chart.style = 10
    chart.y_axis.title = "Option Price (USD)"
    chart.x_axis.title = "Futures Price F"
    chart.height = 12
    chart.width  = 22

    data_ref  = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=len(pF)+1)
    cats_ref  = Reference(ws2, min_col=1, min_row=2, max_row=len(pF)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, "E2")

    wb.save(path)
    print(f"  儲存：{path}")


# ─────────────────────────────────────────────────────────────
# Excel 產生：02 American (BAW)
# ─────────────────────────────────────────────────────────────

def create_baw(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BAW Calculator"
    set_col_widths(ws, [38, 10, 16, 12, 42])

    title_row(ws, 1, "American Futures Option — Barone-Adesi Whaley (BAW) Approximation")
    subtitle_row(ws, 2,
        "Ref: Barone-Adesi & Whaley (1987), J. Finance. "
        "C = c + Ac·(F/F*c)^γc if F < F*c, else F−K")

    # ── 輸入 ──
    sec_header(ws, 4, "▌ INPUT PARAMETERS  （黃色格可修改）")
    inp_row(ws, 5,  "Futures Price",     "F",   100.00, "USD")
    inp_row(ws, 6,  "Strike Price",      "K",   100.00, "USD")
    inp_row(ws, 7,  "Risk-Free Rate",    "r",   0.05,   "decimal")
    inp_row(ws, 8,  "Time to Expiry",    "T",   0.50,   "Years")
    inp_row(ws, 9,  "Volatility",        "σ",   0.20,   "decimal")

    F, K, r, T, s = "C5","C6","C7","C8","C9"

    # ── BAW 參數 ──
    sec_header(ws, 11, "▌ BAW PARAMETERS  (for futures: y = r, n = 0)")
    note_row(ws, 12, "  公式：m = 2rT/σ²，l = 1−e^(−rT)，γ_c = [1+√(1+4m/l)]/2，γ_p = [1−√(1+4m/l)]/2")

    calc_row(ws, 13, "m = 2rT/σ²",           "m",
             f"=2*{r}*{T}/{s}^2",   "")
    calc_row(ws, 14, "l = 1 − e^(−rT)",      "l",
             f"=1-EXP(-{r}*{T})",  "")
    calc_row(ws, 15, "γ_c (call parameter)",  "γ_c",
             "=(1+SQRT(1+4*C13/C14))/2",  "")
    calc_row(ws, 16, "γ_p (put parameter)",   "γ_p",
             "=(1-SQRT(1+4*C13/C14))/2",  "")

    # ── European prices (Black-76) ──
    sec_header(ws, 18, "▌ EUROPEAN COMPONENT  (Black-76, embedded in BAW)")
    calc_row(ws, 19, "d₁  (European)",  "d₁",
             f"=(LN({F}/{K})+0.5*{s}^2*{T})/({s}*SQRT({T}))", "")
    calc_row(ws, 20, "d₂  (European)",  "d₂",
             f"=C19-{s}*SQRT({T})", "")
    calc_row(ws, 21, "European Call c", "c",
             f"=EXP(-{r}*{T})*({F}*NORM.S.DIST(C19,TRUE)-{K}*NORM.S.DIST(C20,TRUE))",
             "USD")
    calc_row(ws, 22, "European Put p",  "p",
             f"=EXP(-{r}*{T})*({K}*NORM.S.DIST(-C20,TRUE)-{F}*NORM.S.DIST(-C19,TRUE))",
             "USD")

    # ── 臨界價格 F* ── (Python 預算值，提供 Goal Seek 驗證)
    res_baw = baw(100, 100, 0.05, 0.50, 0.20)
    sec_header(ws, 24, "▌ CRITICAL PRICE F*  (由 Newton-Raphson 迭代求解)")
    note_row(ws, 25, "  條件：F*_c − K = c(F*_c) + Ac(F*_c)；以下數值由 Python 求解，可用 Goal Seek 驗證")

    # F_star_c 固定值
    fsc = ws.cell(row=26, column=3, value=round(res_baw["fstar_c"], 6))
    apply(ws.cell(row=26, column=1), fill="FFF2CC",
          font=_font(size=10), border=True)
    ws.cell(row=26, column=1).value = "Critical Call Price  F*_c"
    apply(ws.cell(row=26, column=2), fill="FFF2CC",
          font=_font(bold=True, italic=True, color=C_DARK), align=_align("center"), border=True)
    ws.cell(row=26, column=2).value = "F*_c"
    apply(fsc, fill="FFF9E0", font=_font(bold=True, color="C00000"),
          align=_align("right"), num_fmt="#,##0.000000", border=True)
    apply(ws.cell(row=26, column=4), fill="FFF2CC",
          font=_font(size=9, color="666666"), border=True)
    ws.cell(row=26, column=4).value = "USD"

    fsp = ws.cell(row=27, column=3, value=round(res_baw["fstar_p"], 6))
    for col, val, kw in [
        (1, "Critical Put Price   F*_p", dict(fill="FFF2CC", font=_font(size=10))),
        (2, "F*_p", dict(fill="FFF2CC", font=_font(bold=True, italic=True, color=C_DARK), align=_align("center"))),
        (4, "USD",  dict(fill="FFF2CC", font=_font(size=9, color="666666"))),
    ]:
        apply(ws.cell(row=27, column=col), border=True, **kw)
        ws.cell(row=27, column=col).value = val
    apply(fsp, fill="FFF9E0", font=_font(bold=True, color="C00000"),
          align=_align("right"), num_fmt="#,##0.000000", border=True)

    # 邊界條件驗證（應接近 0）
    calc_row(ws, 28, "Call boundary check  [F*_c − K − c(F*_c) − Ac(F*_c)]", "err_c",
             "=C26-C6-EXP(-C7*C8)*(C26*NORM.S.DIST((LN(C26/C6)+0.5*C9^2*C8)/(C9*SQRT(C8)),TRUE)"
             "-C6*NORM.S.DIST((LN(C26/C6)+0.5*C9^2*C8)/(C9*SQRT(C8))-C9*SQRT(C8),TRUE))"
             "-(C26/C15)*(1-EXP(-C7*C8)*NORM.S.DIST((LN(C26/C6)+0.5*C9^2*C8)/(C9*SQRT(C8)),TRUE))",
             "≈ 0")
    note_row(ws, 29, "  ☑ Goal Seek：將 C28 設為 0，調整 C26（F*_c），可重新求解臨界價格")

    # ── A_c, A_p ──
    sec_header(ws, 31, "▌ EARLY EXERCISE COEFFICIENTS")
    calc_row(ws, 32, "Ac = (F*_c / γ_c)·(1 − e^(−rT)·N(d₁(F*_c)))",  "Ac",
             "=(C26/C15)*(1-EXP(-C7*C8)*NORM.S.DIST((LN(C26/C6)+0.5*C9^2*C8)/(C9*SQRT(C8)),TRUE))",
             "")
    calc_row(ws, 33, "Ap = (F*_p / γ_p)·(e^(−rT)·N(−d₁(F*_p)) − 1)", "Ap",
             "=(C27/C16)*(EXP(-C7*C8)*NORM.S.DIST(-((LN(C27/C6)+0.5*C9^2*C8)/(C9*SQRT(C8))),TRUE)-1)",
             "")

    # ── American Prices ──
    sec_header(ws, 35, "▌ AMERICAN OPTION PRICES")
    calc_row(ws, 36, "American Call  C(F,T)  [if F < F*_c: c + Ac·(F/F*_c)^γ_c]", "C_Am",
             f"=IF({F}>=C26,{F}-{K},C21+C32*({F}/C26)^C15)",
             "USD", is_out=True)
    note_row(ws, 37, "  ※ 若 F ≥ F*_c，則立即行使（Immediate Exercise Value = F − K）")
    calc_row(ws, 38, "American Put   P(F,T)  [if F > F*_p: p + Ap·(F/F*_p)^γ_p]", "P_Am",
             f"=IF({F}<=C27,{K}-{F},C22+C33*({F}/C27)^C16)",
             "USD", is_out=True)
    note_row(ws, 39, "  ※ 若 F ≤ F*_p，則立即行使（Immediate Exercise Value = K − F）")
    calc_row(ws, 40, "Early Exercise Premium — Call  (C_Am − c)", "EEP_c",
             "=C36-C21", "USD")
    calc_row(ws, 41, "Early Exercise Premium — Put   (P_Am − p)", "EEP_p",
             "=C38-C22", "USD")

    # ── 敏感度表（Python 預算） ──
    sec_header(ws, 43, "▌ SENSITIVITY TABLE — American vs European (K=100, r=5%, T=0.5yr, σ=20%)")
    pF = np.arange(70, 135, 5)
    h = ["Futures Price F"] + [f"{f:.0f}" for f in pF]
    for j, v in enumerate(h):
        c = ws.cell(row=44, column=j+1, value=v)
        apply(c, fill=C_DARK, font=_font(bold=True, color=C_WHITE, size=9),
              align=_align("center"))

    rows = {"Eur. Call": [], "Am. Call": [], "EEP (Call)": [],
            "Eur. Put":  [], "Am. Put":  [], "EEP (Put)":  []}
    for f in pF:
        e = black76(f, 100, 0.05, 0.5, 0.2)
        a = baw(f, 100, 0.05, 0.5, 0.2)
        rows["Eur. Call"].append(e["call"])
        rows["Am. Call"].append(a["call"])
        rows["EEP (Call)"].append(a["call"] - e["call"])
        rows["Eur. Put"].append(e["put"])
        rows["Am. Put"].append(a["put"])
        rows["EEP (Put)"].append(a["put"] - e["put"])

    for ri, (lbl, vals) in enumerate(rows.items(), start=45):
        ws.cell(row=ri, column=1, value=lbl)
        apply(ws.cell(row=ri, column=1), fill=C_CALC, font=_font(size=9))
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=ri, column=j, value=round(v, 4))
            apply(c, fill="FFFFFF" if ri%2==0 else C_CALC,
                  font=_font(size=9), num_fmt="#,##0.0000")

    wb.save(path)
    print(f"  儲存：{path}")


# ─────────────────────────────────────────────────────────────
# Excel 產生：03 Quanto
# ─────────────────────────────────────────────────────────────

def create_quanto(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quanto Calculator"
    set_col_widths(ws, [38, 10, 16, 14, 44])

    title_row(ws, 1, "Quanto Futures Option  （固定匯率 Quanto）")
    subtitle_row(ws, 2,
        "Payoff（payoff currency）= (1/X̄)·max(F_T − K, 0)  "
        "Drift adjustment: F_adj = F·exp(−ρ·σ·σ_X·T)")

    sec_header(ws, 4, "▌ INPUT PARAMETERS  （黃色格可修改）")
    inp_row(ws, 5,  "Futures Price (underlier currency)", "F",     100.00, "USD")
    inp_row(ws, 6,  "Strike (underlier currency)",        "K",     100.00, "USD")
    inp_row(ws, 7,  "Payoff Currency Risk-Free Rate",     "r_d",   0.04,   "decimal", "歐元等 payoff 幣別利率")
    inp_row(ws, 8,  "Time to Expiry",                    "T",     0.50,   "Years")
    inp_row(ws, 9,  "Underlying Volatility",              "σ",     0.20,   "decimal")
    inp_row(ws, 10, "FX Rate Volatility",                "σ_X",   0.10,   "decimal", "F/X FX vol")
    inp_row(ws, 11, "Correlation  (ρ: F return vs FX)",  "ρ",     0.30,   "",        "−1 ≤ ρ ≤ 1")
    inp_row(ws, 12, "Fixed Exchange Rate",               "X̄",    1.20,   "USD/EUR", "Quanto 固定匯率，X̄=1 表示 1:1")

    F,K,r,T,s,sx,rho,Xb = "C5","C6","C7","C8","C9","C10","C11","C12"

    sec_header(ws, 14, "▌ DRIFT ADJUSTMENT  — Quanto Correction")
    note_row(ws, 15, "  在 payoff 幣別測度下，F 的漂移項須加上相關性調整：−ρ·σ·σ_X·T")
    calc_row(ws, 16, "Correlation Adjustment  −ρ·σ·σ_X·T",      "adj",
             f"=-{rho}*{s}*{sx}*{T}", "")
    calc_row(ws, 17, "Drift-Adjusted Forward  F_adj = F·exp(adj)", "F_adj",
             f"={F}*EXP(C16)", "USD")

    sec_header(ws, 19, "▌ BLACK-76 ON ADJUSTED FORWARD")
    calc_row(ws, 20, "d₁  (using F_adj)", "d₁",
             f"=(LN(C17/{K})+0.5*{s}^2*{T})/({s}*SQRT({T}))", "")
    calc_row(ws, 21, "d₂",               "d₂",
             f"=C20-{s}*SQRT({T})", "")
    calc_row(ws, 22, "Discount Factor",  "e^(−r_d·T)",
             f"=EXP(-{r}*{T})", "")

    sec_header(ws, 24, "▌ QUANTO OPTION PRICES  （in payoff currency）")
    calc_row(ws, 25, "Quanto Call  = (1/X̄)·e^(−r_d·T)·[F_adj·N(d₁) − K·N(d₂)]", "C_Q",
             f"=(1/{Xb})*C22*(C17*NORM.S.DIST(C20,TRUE)-{K}*NORM.S.DIST(C21,TRUE))",
             "EUR", is_out=True)
    calc_row(ws, 26, "Quanto Put   = (1/X̄)·e^(−r_d·T)·[K·N(−d₂) − F_adj·N(−d₁)]", "P_Q",
             f"=(1/{Xb})*C22*({K}*NORM.S.DIST(-C21,TRUE)-C17*NORM.S.DIST(-C20,TRUE))",
             "EUR", is_out=True)
    calc_row(ws, 27, "Put-Call Parity Check",  "✓",
             f"=C25-C26-C22*(C17-{K})/{Xb}", "≈ 0")
    note_row(ws, 28, "  ☑ PCP：C_Q − P_Q = (1/X̄)·e^(−r_d·T)·(F_adj − K)，應接近 0")

    sec_header(ws, 30, "▌ QUANTO GREEKS  (per unit)")
    calc_row(ws, 31, "Quanto Delta — Call  Δ_c = (1/X̄)·e^(−r_d·T)·N(d₁)", "Δ_c",
             f"=(1/{Xb})*C22*NORM.S.DIST(C20,TRUE)", "")
    calc_row(ws, 32, "Quanto Vega  ν (per 1% σ)", "ν",
             f"=(1/{Xb})*C17*C22*NORM.S.DIST(C20,FALSE)*SQRT({T})*0.01", "EUR/1%σ")

    # 敏感度表格
    sec_header(ws, 34, "▌ SENSITIVITY TABLE — Quanto Price vs Futures Price & Correlation")
    pF = np.arange(70, 135, 5)
    rhos = [-0.5, -0.3, 0.0, 0.3, 0.5]
    hdr = ["F \\ ρ"] + [f"ρ={p:+.1f}" for p in rhos]
    for j, v in enumerate(hdr):
        c = ws.cell(row=35, column=j+1, value=v)
        apply(c, fill=C_DARK, font=_font(bold=True, color=C_WHITE, size=9),
              align=_align("center"))
    for ri, f in enumerate(pF, start=36):
        ws.cell(row=ri, column=1, value=float(f))
        apply(ws.cell(row=ri, column=1), fill=C_CALC, font=_font(size=9, bold=True))
        for j, rh in enumerate(rhos, start=2):
            v = quanto(f, 100, 0.04, 0.5, 0.20, 0.10, rh, 1.20, "call")["price"]
            c = ws.cell(row=ri, column=j, value=round(v, 4))
            apply(c, fill="FFFFFF" if ri%2==0 else C_CALC,
                  font=_font(size=9), num_fmt="#,##0.0000")

    wb.save(path)
    print(f"  儲存：{path}")


# ─────────────────────────────────────────────────────────────
# Excel 產生：04 Cross-Currency
# ─────────────────────────────────────────────────────────────

def create_cross_currency(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cross-Currency Calculator"
    set_col_widths(ws, [40, 10, 16, 14, 44])

    title_row(ws, 1, "Cross-Currency Futures Option  （跨幣種選擇權）")
    subtitle_row(ws, 2,
        "Payoff（payoff ccy）= max(F_T/X_T − K, 0)，"
        "Effective Vol σ_Z = √(σ²+σ_X²+2ρσσ_X)，Z₀ = F/X₀")

    sec_header(ws, 4, "▌ INPUT PARAMETERS  （黃色格可修改）")
    inp_row(ws, 5,  "Futures Price (underlier currency)",  "F",    4200.0, "USD",    "如 S&P 500 期貨")
    inp_row(ws, 6,  "Strike (payoff currency)",            "K",    3600.0, "EUR",    "執行價為 payoff 幣種")
    inp_row(ws, 7,  "Payoff Currency Risk-Free Rate",      "r_d",  0.04,   "decimal","EUR 利率")
    inp_row(ws, 8,  "Time to Expiry",                     "T",    0.50,   "Years")
    inp_row(ws, 9,  "Underlying Volatility",               "σ",    0.20,   "decimal")
    inp_row(ws, 10, "FX Rate Volatility  σ_X",            "σ_X",  0.10,   "decimal","USD/EUR FX vol")
    inp_row(ws, 11, "Correlation  ρ(F, 1/X)",             "ρ",    0.30,   "",       "F 與 1/X（ρ 同號代表強化）")
    inp_row(ws, 12, "Spot FX Rate  X₀ (payoff/underlier)","X₀",   0.92,   "EUR/USD","今日即期匯率")

    F,K,r,T,s,sx,rho,X0 = "C5","C6","C7","C8","C9","C10","C11","C12"

    sec_header(ws, 14, "▌ EFFECTIVE VOLATILITY  σ_Z = √(σ²+σ_X²+2ρσσ_X)")
    note_row(ws, 15,
        "  期貨價格（underlier ccy）÷ FX rate = 期貨價格（payoff ccy）；"
        "其波動度由 σ、σ_X、ρ 合成")
    calc_row(ws, 16, "Effective Volatility  σ_Z = √(σ²+σ_X²+2ρ·σ·σ_X)", "σ_Z",
             f"=SQRT({s}^2+{sx}^2+2*{rho}*{s}*{sx})", "decimal")
    calc_row(ws, 17, "Futures Price in Payoff Ccy  Z₀ = F / X₀",          "Z₀",
             f"={F}/{X0}", "EUR")

    sec_header(ws, 19, "▌ BLACK-76 ON Z₀  (payoff currency)")
    calc_row(ws, 20, "d₁  (using Z₀ & σ_Z)", "d₁",
             f"=(LN(C17/{K})+0.5*C16^2*{T})/(C16*SQRT({T}))", "")
    calc_row(ws, 21, "d₂", "d₂",
             f"=C20-C16*SQRT({T})", "")
    calc_row(ws, 22, "Discount Factor",  "e^(−r_d·T)",
             f"=EXP(-{r}*{T})", "")

    sec_header(ws, 24, "▌ CROSS-CURRENCY OPTION PRICES  （in payoff currency）")
    calc_row(ws, 25,
             "Cross-Ccy Call  = e^(−r_d·T)·[Z₀·N(d₁) − K·N(d₂)]", "C_X",
             f"=C22*(C17*NORM.S.DIST(C20,TRUE)-{K}*NORM.S.DIST(C21,TRUE))",
             "EUR", is_out=True)
    calc_row(ws, 26,
             "Cross-Ccy Put   = e^(−r_d·T)·[K·N(−d₂) − Z₀·N(−d₁)]", "P_X",
             f"=C22*({K}*NORM.S.DIST(-C21,TRUE)-C17*NORM.S.DIST(-C20,TRUE))",
             "EUR", is_out=True)
    calc_row(ws, 27, "Put-Call Parity Check  [C_X − P_X − e^(−r_d·T)(Z₀−K)]", "✓",
             f"=C25-C26-C22*(C17-{K})", "≈ 0")

    sec_header(ws, 29, "▌ CROSS-CURRENCY GREEKS  (per unit)")
    calc_row(ws, 30, "Delta in Payoff Ccy  ∂C_X/∂Z₀ = e^(−r_d·T)·N(d₁)", "Δ_Z",
             "=C22*NORM.S.DIST(C20,TRUE)", "")
    calc_row(ws, 31, "Delta in Underlier Ccy  ∂C_X/∂F = Δ_Z / X₀",        "Δ_F",
             f"=C30/{X0}", "per USD")
    calc_row(ws, 32, "Effective Vega  (per 1% σ_Z)",                        "ν_Z",
             f"=C17*C22*NORM.S.DIST(C20,FALSE)*SQRT({T})*0.01", "EUR/1%σ")

    # 敏感度表格：不同 FX Vol 與 Correlation 對有效波動度的影響
    sec_header(ws, 34, "▌ EFFECTIVE VOL TABLE — σ_X vs ρ  (σ=20%)")
    sx_vals = [0.05, 0.10, 0.15, 0.20, 0.25]
    rho_vals = [-0.5, -0.3, 0.0, 0.3, 0.5]
    hdr = ["σ_X \\ ρ"] + [f"ρ={p:+.1f}" for p in rho_vals]
    for j, v in enumerate(hdr):
        c = ws.cell(row=35, column=j+1, value=v)
        apply(c, fill=C_DARK, font=_font(bold=True, color=C_WHITE, size=9),
              align=_align("center"))
    for ri, sxv in enumerate(sx_vals, start=36):
        ws.cell(row=ri, column=1, value=f"{sxv*100:.0f}%")
        apply(ws.cell(row=ri, column=1), fill=C_CALC, font=_font(size=9, bold=True))
        for j, rv in enumerate(rho_vals, start=2):
            sv = np.sqrt(0.20**2 + sxv**2 + 2*rv*0.20*sxv)
            c = ws.cell(row=ri, column=j, value=round(sv, 4))
            apply(c, fill="FFFFFF" if ri%2==0 else C_CALC,
                  font=_font(size=9), num_fmt="0.00%")

    # 價格敏感度表格
    sec_header(ws, 42, "▌ PRICE TABLE — Cross-Ccy Call vs Z₀ and σ_Z")
    pZ = np.arange(2800, 4800, 200)
    svs = [0.15, 0.20, 0.25, 0.30]
    hdr2 = ["Z₀ \\ σ_Z"] + [f"σ_Z={v*100:.0f}%" for v in svs]
    for j, v in enumerate(hdr2):
        c = ws.cell(row=43, column=j+1, value=v)
        apply(c, fill=C_DARK, font=_font(bold=True, color=C_WHITE, size=9),
              align=_align("center"))
    for ri, z in enumerate(pZ, start=44):
        ws.cell(row=ri, column=1, value=float(z))
        apply(ws.cell(row=ri, column=1), fill=C_CALC, font=_font(size=9, bold=True))
        for j, sv in enumerate(svs, start=2):
            r0 = 0.04; T0 = 0.5; K0 = 3600.0
            if z <= 0 or K0 <= 0 or sv <= 0:
                continue
            d1 = (np.log(z/K0) + 0.5*sv**2*T0)/(sv*np.sqrt(T0))
            d2 = d1 - sv*np.sqrt(T0)
            price = np.exp(-r0*T0)*(z*norm.cdf(d1) - K0*norm.cdf(d2))
            c = ws.cell(row=ri, column=j, value=round(price, 2))
            apply(c, fill="FFFFFF" if ri%2==0 else C_CALC,
                  font=_font(size=9), num_fmt="#,##0.00")

    wb.save(path)
    print(f"  儲存：{path}")


# ─────────────────────────────────────────────────────────────
# 驗證（印出數值，確認 Python 函式正確）
# ─────────────────────────────────────────────────────────────

def verify():
    print("\n=== 驗證計算結果 ===")
    # European
    e = black76(100, 100, 0.05, 0.5, 0.20)
    print(f"Black-76  Call={e['call']:.4f}  Put={e['put']:.4f}  "
          f"PCP={e['call']-e['put'] - np.exp(-0.05*0.5)*(100-100):.2e}")

    # American
    a = baw(100, 100, 0.05, 0.5, 0.20)
    print(f"BAW       Call={a['call']:.4f}  Put={a['put']:.4f}  "
          f"F*_c={a['fstar_c']:.4f}  F*_p={a['fstar_p']:.4f}")
    print(f"BAW EEP   Call={a['call']-e['call']:.4f}  Put={a['put']-e['put']:.4f}")

    # Quanto
    q = quanto(100, 100, 0.04, 0.5, 0.20, 0.10, 0.30, 1.20, "call")
    print(f"Quanto    Call={q['price']:.4f}  F_adj={q['F_adj']:.4f}")

    # Cross-currency
    cc = cross_currency(4200, 3600, 0.04, 0.5, 0.20, 0.10, 0.30, 0.92, "call")
    print(f"Cross-Ccy Call={cc['price']:.4f}  Z0={cc['Z0']:.4f}  σ_eff={cc['sigma_eff']:.4f}")


# ─────────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("開始產生 Excel 計算範例...")

    verify()
    print()

    paths = {
        "01_European_Futures_Option.xlsx":      create_european,
        "02_American_Futures_Option_BAW.xlsx":  create_baw,
        "03_Quanto_Futures_Option.xlsx":        create_quanto,
        "04_CrossCurrency_Futures_Option.xlsx": create_cross_currency,
    }
    for fname, func in paths.items():
        fpath = os.path.join(OUTPUT_DIR, fname)
        print(f"產生 {fname}...")
        func(fpath)

    print("\n全部完成！檔案位於：", OUTPUT_DIR)
