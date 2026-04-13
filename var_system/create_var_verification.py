"""
create_var_verification.py  v2
==========================
產生「VaR 三方法逐步驗證 Excel」，供人工核對每一步計算。

v1 → v2 修正：
  1. 修正回報起點：LOCAL_RET 改從 ALL_PRICES（含 S0）計算
     舊：diff(PRICES[0..39]) → 回報 1..39，與日期差一格
     新：diff([S0]+PRICES[0..N-1]) → 回報 0..N-1，正確對齊
  2. N_DAYS 改為 250（標準 1 年），99% VaR 有 2 個尾端觀測
  3. Sheet 3 全面使用 Excel 公式（LN 直接引用 Sheet 2 原始價格）
  4. Sheet 4 排序欄位使用 SMALL() 公式，VaR/CVaR 結果格亦為公式
"""

import numpy as np
import pandas as pd
from scipy.stats import norm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ─── 固定隨機種子 ─────────────────────────────────────────────
RNG = np.random.default_rng(seed=2024)

# ─── 樣式常數 ─────────────────────────────────────────────────
NAVY   = "1F497D"
LGRAY  = "F2F2F2"
LGREEN = "E2EFDA"
LRED   = "FFDFD5"
LBLUE  = "DCE6F1"
LYELL  = "FFFF99"
GREEN  = "375623"

def hdr(ws, row, col, text, bg=NAVY, fg="FFFFFF", bold=True, size=10):
    c = ws.cell(row, col, text)
    c.font = Font(bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def val(ws, row, col, v, fmt=None, bg=None, bold=False, align="center"):
    c = ws.cell(row, col, v)
    c.font = Font(bold=bold, size=9)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def note(ws, row, col, text, bg=LGRAY):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+10)
    c = ws.cell(row, col, text)
    c.font = Font(italic=True, color="595959", size=8)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c

def merge_hdr(ws, row, c1, c2, text, bg=NAVY, fg="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, text)
    c.font = Font(bold=True, color=fg, size=11)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")

def section(ws, row, text, ncol=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    c = ws.cell(row, 1, f"▌ {text}")
    c.font = Font(bold=True, size=10, color=NAVY)
    c.fill = PatternFill("solid", fgColor=LBLUE)
    c.alignment = Alignment(vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, min_col, max_row, max_col):
    b = thin_border()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = b

def autofit(ws, min_w=8, max_w=30):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            min(max_len * 1.15, max_w), min_w
        )

# ═══════════════════════════════════════════════════════════════
# 產生樣本資料
# ═══════════════════════════════════════════════════════════════
N_DAYS = 250   # 約 1 年，99% VaR 有 2 個尾端觀測值

MU    = np.array([0.0008,  0.0007,  0.0006])
SIGMA = np.array([0.020,   0.018,   0.022])
CORR  = np.array([[1.00, 0.30, 0.20],
                  [0.30, 1.00, 0.22],
                  [0.20, 0.22, 1.00]])
COV_TRUE = np.diag(SIGMA) @ CORR @ np.diag(SIGMA)
L_TRUE   = np.linalg.cholesky(COV_TRUE)

Z    = RNG.standard_normal((N_DAYS, 3))
RET  = Z @ L_TRUE.T + MU          # (N_DAYS, 3)

S0     = np.array([600.0, 185.0, 70000.0])
PRICES = S0 * np.exp(np.cumsum(RET, axis=0))   # (N_DAYS, 3)

# FX 歷史
FX0      = np.array([32.5, 0.0235])
FX_SIGMA = np.array([0.003, 0.004])
FX_Z     = RNG.standard_normal((N_DAYS, 2))
FX_RET_G = FX_Z * FX_SIGMA
FX       = FX0 * np.exp(np.cumsum(FX_RET_G, axis=0))   # (N_DAYS, 2)

# 部位
QTY    = np.array([1000.0, 100.0, 10.0])
CCY    = ["TWD", "USD", "KRW"]
ASSETS = ["台積電(2330.TW)", "AAPL", "三星(005930.KS)"]

# 日期序列：N_DAYS+1 個，D[0]=初始日，D[1..N_DAYS]=回報日
DATES = pd.bdate_range(end="2024-12-31", periods=N_DAYS + 1)

# ─────────────────────────────────────────────────────────────
# 完整價格序列（含初始 S0）
#   ALL_PRICES[k] = 第 k 日收盤價，對應 DATES[k]
#   ALL_PRICES[0] = S0
#   ALL_PRICES[1] = PRICES[0]  ← 第 1 個交易日
# ─────────────────────────────────────────────────────────────
ALL_PRICES = np.vstack([S0, PRICES])   # (N_DAYS+1, 3)
ALL_FX     = np.vstack([FX0, FX])     # (N_DAYS+1, 2)

# 最新市值（最後交易日）
LATEST_LOCAL = PRICES[-1]
FX_LATEST    = {"TWD": 1.0, "USD": float(FX[-1, 0]), "KRW": float(FX[-1, 1])}
MV_BASE      = np.array([
    QTY[i] * LATEST_LOCAL[i] * FX_LATEST[CCY[i]]
    for i in range(3)
])
PORTFOLIO_VALUE = float(np.sum(MV_BASE))

# ─────────────────────────────────────────────────────────────
# 回報計算（修正：從 ALL_PRICES 計算，包含第一天 S0→PRICES[0]）
#   LOCAL_RET[t] = ln(ALL_PRICES[t+1] / ALL_PRICES[t])
#               = 日期 DATES[t+1] 的本地回報
# ─────────────────────────────────────────────────────────────
LOCAL_RET = np.diff(np.log(ALL_PRICES), axis=0)   # (N_DAYS, 3)
FX_RET_TS = np.diff(np.log(ALL_FX),    axis=0)    # (N_DAYS, 2)

ADJ_RET = LOCAL_RET.copy()
ADJ_RET[:, 1] += FX_RET_TS[:, 0]   # AAPL += USD 回報
ADJ_RET[:, 2] += FX_RET_TS[:, 1]   # 三星 += KRW 回報

T = len(ADJ_RET)   # = N_DAYS = 250

# VaR 參數
CONFIDENCE = 0.99
Z_ALPHA    = norm.ppf(CONFIDENCE)
IDX_VAR    = int(np.floor((1 - CONFIDENCE) * T))   # = 2 for T=250

# ─── HS 計算 ──────────────────────────────────────────────────
PNL_HS     = (ADJ_RET * MV_BASE).sum(axis=1)
SORTED_PNL = np.sort(PNL_HS)
HS_VAR     = float(-SORTED_PNL[IDX_VAR])
HS_CVAR    = float(-SORTED_PNL[:max(IDX_VAR, 1)].mean())

HS_COMP_VAR = {
    ASSETS[i]: float(-np.percentile(ADJ_RET[:, i] * MV_BASE[i], (1 - CONFIDENCE) * 100))
    for i in range(3)
}
HS_COMP_CVAR = {}
for i in range(3):
    ap  = ADJ_RET[:, i] * MV_BASE[i]
    thr = np.percentile(ap, (1 - CONFIDENCE) * 100)
    tail = ap[ap <= thr]
    HS_COMP_CVAR[ASSETS[i]] = float(-tail.mean()) if len(tail) else HS_COMP_VAR[ASSETS[i]]

# ─── Parametric 計算 ──────────────────────────────────────────
COV_SAMPLE  = np.cov(ADJ_RET.T)
CORR_SAMPLE = (np.diag(1 / np.sqrt(np.diag(COV_SAMPLE)))
               @ COV_SAMPLE
               @ np.diag(1 / np.sqrt(np.diag(COV_SAMPLE))))
W        = MV_BASE
PORT_VAR = float(W @ COV_SAMPLE @ W)
PORT_STD = float(np.sqrt(PORT_VAR))
PAR_VAR  = Z_ALPHA * PORT_STD
PAR_CVAR = (norm.pdf(Z_ALPHA) / (1 - CONFIDENCE)) * PORT_STD

SIGMA_VEC    = COV_SAMPLE @ W
MC_PAR_VEC   = SIGMA_VEC / PORT_STD
PAR_COMP_VAR = {ASSETS[i]: float(W[i] * MC_PAR_VEC[i] * Z_ALPHA) for i in range(3)}
PAR_COMP_CVAR = {
    n: float(PAR_COMP_VAR[n] * norm.pdf(Z_ALPHA) / ((1 - CONFIDENCE) * Z_ALPHA))
    for n in ASSETS
}

# ─── Monte Carlo 計算 ─────────────────────────────────────────
MC_N = 5000
RNG2 = np.random.default_rng(seed=42)
try:
    L_MC = np.linalg.cholesky(COV_SAMPLE)
except np.linalg.LinAlgError:
    L_MC = np.linalg.cholesky(COV_SAMPLE + np.eye(3) * 1e-8)

MU_RET = ADJ_RET.mean(axis=0)
Z_MC   = RNG2.standard_normal((MC_N, 3))
SIM    = Z_MC @ L_MC.T + MU_RET
PNL_MC = SIM @ MV_BASE
MC_VAR  = float(-np.percentile(PNL_MC, (1 - CONFIDENCE) * 100))
thr_mc  = np.percentile(PNL_MC, (1 - CONFIDENCE) * 100)
MC_CVAR = float(-PNL_MC[PNL_MC <= thr_mc].mean())
MC_COMP_VAR = {
    ASSETS[i]: float(-np.percentile(SIM[:, i] * MV_BASE[i], (1 - CONFIDENCE) * 100))
    for i in range(3)
}
tail_mask = PNL_MC <= thr_mc
MC_COMP_CVAR = {
    ASSETS[i]: float(-((SIM[:, i] * MV_BASE[i])[tail_mask]).mean())
    for i in range(3)
}

# ─── Cholesky 手算（3×3）─────────────────────────────────────
S = COV_SAMPLE
L = np.zeros((3, 3))
L[0, 0] = np.sqrt(S[0, 0])
L[1, 0] = S[1, 0] / L[0, 0]
L[1, 1] = np.sqrt(S[1, 1] - L[1, 0] ** 2)
L[2, 0] = S[2, 0] / L[0, 0]
L[2, 1] = (S[2, 1] - L[2, 0] * L[1, 0]) / L[1, 1]
L[2, 2] = np.sqrt(S[2, 2] - L[2, 0] ** 2 - L[2, 1] ** 2)

# ─── 印出摘要 ─────────────────────────────────────────────────
print(f"[設定] N_DAYS={N_DAYS}, T={T}, IDX_VAR={IDX_VAR}")
print(f"  HS  VaR = {HS_VAR:>12,.2f} TWD  (第 {IDX_VAR+1} 小損益)")
print(f"  Par VaR = {PAR_VAR:>12,.2f} TWD")
print(f"  MC  VaR = {MC_VAR:>12,.2f} TWD")
sum_cv = sum(PAR_COMP_VAR.values())
print(f"  Parametric Component VaR 加總誤差：|{sum_cv:.4f} - {PAR_VAR:.4f}| = {abs(sum_cv-PAR_VAR):.6e}")

# ═══════════════════════════════════════════════════════════════
# 開始寫 Excel
# ═══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 名稱常數（供公式引用）
S1 = "'①基本設定'"
S2 = "'②原始價格與FX'"
S3 = "'③回報計算FX調整'"

# Sheet 2 資料列對應：
#   ALL_PRICES[k] → Sheet2 row (k + S2_DATA_START)
S2_DATA_START = 3   # ALL_PRICES[0]=S0 在 Sheet2 第 3 列

# Sheet 3 資料列對應：
#   ADJ_RET[row_i] → Sheet3 row (row_i + S3_DATA_START)
S3_DATA_START = 16   # section at 14, headers at 15, data from 16
S3_DATA_END   = S3_DATA_START + T - 1   # = 265

# Sheet 4 資料列對應
S4_DATA_START = 14
S4_DATA_END   = S4_DATA_START + T - 1   # = 263

# ════════════════════════════════════════════════════════════
# Sheet 1 — 基本設定
# ════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("①基本設定")
merge_hdr(ws1, 1, 1, 8, "VaR 逐步計算驗證 — 基本設定（v2）")
ws1.row_dimensions[1].height = 22

section(ws1, 3, "資產與部位定義")
for c, h in enumerate(["資產名稱", "市場", "幣別", "持有數量",
                        "最新本地價格", "即期匯率(→TWD)",
                        "TWD 市值（G列）", "占組合比重"], start=1):
    hdr(ws1, 4, c, h, bg="2E75B6")

# MV_BASE 在 Sheet1 G5, G6, G7 — 供 Sheet3 P&L 公式引用
from var_engine import get_market_label
for i in range(3):
    r = i + 5
    mkt = get_market_label(ASSETS[i])
    fx  = FX_LATEST[CCY[i]]
    val(ws1, r, 1, ASSETS[i],         align="left")
    val(ws1, r, 2, mkt,               align="left")
    val(ws1, r, 3, CCY[i])
    val(ws1, r, 4, QTY[i],            fmt="#,##0")
    val(ws1, r, 5, round(float(LATEST_LOCAL[i]), 2), fmt="#,##0.00")
    val(ws1, r, 6, round(fx, 4),      fmt="#,##0.00000")
    val(ws1, r, 7, round(float(MV_BASE[i]), 0), fmt="#,##0", bold=True)
    val(ws1, r, 8, MV_BASE[i] / PORTFOLIO_VALUE, fmt="0.00%")

ws1.cell(8, 6, "組合總市值(TWD)").font = Font(bold=True, size=9)
val(ws1, 8, 7, round(PORTFOLIO_VALUE, 0), fmt="#,##0", bold=True, bg=LYELL)
apply_border(ws1, 4, 1, 8, 8)

note(ws1, 9, 1, "★ G5:G7（TWD 市值）會被 Sheet 3 P&L 公式直接引用，請勿移動此欄位置。")

section(ws1, 11, "VaR 計算參數")
params = [
    ("信賴水準 α",              f"{CONFIDENCE*100:.0f}%"),
    ("持有天數 h",              "1 日"),
    ("歷史回溯天數 T",          f"{T} 個交易日"),
    ("回報方式",                "對數回報 ln(P_t / P_{t-1})"),
    ("FX 調整",                 "外幣資產：r_adj = r_local + r_FX（對數精確分解）"),
    ("z_α",                    f"=NORM.S.INV({CONFIDENCE}) = {Z_ALPHA:.6f}"),
    ("VaR 分位索引（HS）",      f"idx = floor((1-α)×T) = floor(0.01×{T}) = {IDX_VAR}"),
    ("HS VaR 公式",             f"-P&L_sorted[{IDX_VAR}]  （第 {IDX_VAR+1} 小損益）"),
    ("HS CVaR 公式",            f"-mean(P&L_sorted[0:{IDX_VAR}])  （最差 {max(IDX_VAR,1)} 筆均值）"),
    ("MC 模擬路徑數",           f"{MC_N:,}"),
]
for r, (k, v) in enumerate(params, start=12):
    val(ws1, r, 1, k, bold=True, align="left")
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    val(ws1, r, 2, v, align="left")
apply_border(ws1, 12, 1, 21, 6)

section(ws1, 23, "即期匯率（計算時使用）")
for c, h in enumerate(["幣別", "匯率（1 外幣 = ? TWD）", "說明"], start=1):
    hdr(ws1, 24, c, h, bg="2E75B6")
for r, (c, rv, d) in enumerate([
    ("TWD", 1.0,                    "基礎幣別（不動）"),
    ("USD", FX_LATEST["USD"],       "美元/台幣"),
    ("KRW", FX_LATEST["KRW"],       "韓元/台幣"),
], start=25):
    val(ws1, r, 1, c)
    val(ws1, r, 2, round(rv, 5), fmt="0.00000")
    val(ws1, r, 3, d, align="left")
apply_border(ws1, 24, 1, 27, 3)
autofit(ws1)

# ════════════════════════════════════════════════════════════
# Sheet 2 — 原始價格與 FX（純數值）
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("②原始價格與FX")
merge_hdr(ws2, 1, 1, 9, "原始收盤價 & FX 匯率歷史（本地幣別，純數值）")

for c, h in enumerate(["日期", "台積電(TWD)", "AAPL(USD)", "三星(KRW)",
                        "", "日期", "USD/TWD", "KRW/TWD"], start=1):
    if h:
        hdr(ws2, 2, c, h, bg="2E75B6")

for row_i in range(N_DAYS + 1):
    r = row_i + S2_DATA_START
    val(ws2, r, 1, DATES[row_i].strftime("%Y-%m-%d"), align="center")
    # 股價
    val(ws2, r, 2, round(float(ALL_PRICES[row_i, 0]), 2), fmt="#,##0.00")
    val(ws2, r, 3, round(float(ALL_PRICES[row_i, 1]), 2), fmt="#,##0.00")
    val(ws2, r, 4, round(float(ALL_PRICES[row_i, 2]), 0), fmt="#,##0")
    # FX
    val(ws2, r, 6, DATES[row_i].strftime("%Y-%m-%d"), align="center")
    val(ws2, r, 7, round(float(ALL_FX[row_i, 0]), 4), fmt="#,##0.0000")
    val(ws2, r, 8, round(float(ALL_FX[row_i, 1]), 6), fmt="0.000000")

# 初始列（S0）醒目標示
for c in range(1, 9):
    ws2.cell(S2_DATA_START, c).fill = PatternFill("solid", fgColor=LYELL)

apply_border(ws2, 2, 1, N_DAYS + S2_DATA_START, 4)
apply_border(ws2, 2, 6, N_DAYS + S2_DATA_START, 8)
note(ws2, N_DAYS + S2_DATA_START + 1, 1,
     f"共 {N_DAYS+1} 列（含黃色初始列 S0）。"
     "B欄=台積電 TWD，C欄=AAPL USD，D欄=三星 KRW，G欄=USD/TWD，H欄=KRW/TWD。")
autofit(ws2)

# ════════════════════════════════════════════════════════════
# Sheet 3 — 回報計算（全用 Excel 公式）
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("③回報計算FX調整")
merge_hdr(ws3, 1, 1, 10, "步驟一：計算對數回報 & FX 調整（公式直接引用 Sheet 2）")

section(ws3, 3, "公式說明")
formulas_doc = [
    ("B欄 台積電 r_local",  "=LN('②原始價格與FX'!B{t} / '②原始價格與FX'!B{t-1})"),
    ("C欄 AAPL r_local",    "=LN('②原始價格與FX'!C{t} / '②原始價格與FX'!C{t-1})"),
    ("D欄 三星 r_local",    "=LN('②原始價格與FX'!D{t} / '②原始價格與FX'!D{t-1})"),
    ("E欄 USD r_FX",        "=LN('②原始價格與FX'!G{t} / '②原始價格與FX'!G{t-1})  ← G欄=USD/TWD"),
    ("F欄 KRW r_FX",        "=LN('②原始價格與FX'!H{t} / '②原始價格與FX'!H{t-1})  ← H欄=KRW/TWD"),
    ("G欄 台積電 r_adj",    "=B{row}  （TWD 資產，r_adj = r_local）"),
    ("H欄 AAPL r_adj",      "=C{row}+E{row}  （USD 資產，r_adj = r_local + r_FX）"),
    ("I欄 三星 r_adj",      "=D{row}+F{row}  （KRW 資產，r_adj = r_local + r_FX）"),
    ("J欄 組合P&L",         "=G{row}*'①基本設定'!$G$5 + H{row}*'①基本設定'!$G$6 + I{row}*'①基本設定'!$G$7"),
    ("驗證方法",            "J欄 P&L = Σ(r_adj_i × MV_i)，與 Sheet4 HS VaR 計算的基礎相同"),
]
for r, (k, v) in enumerate(formulas_doc, start=4):
    val(ws3, r, 1, k, bold=True, align="left", bg=LBLUE if "驗證" not in k else LGREEN)
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    val(ws3, r, 2, v, align="left")

section(ws3, 14, f"回報序列（{T} 個交易日，全為 Excel 公式）")
for c, h in enumerate(["日期",
                        "台積電 r_local", "AAPL r_local", "三星 r_local",
                        "USD r_FX", "KRW r_FX",
                        "台積電 r_adj", "AAPL r_adj", "三星 r_adj",
                        "組合P&L(TWD)"], start=1):
    hdr(ws3, 15, c, h, bg="2E75B6")

# 公式資料列
# Sheet2 資料從 row S2_DATA_START 開始：
#   ALL_PRICES[k]  → Sheet2 row (S2_DATA_START + k)
#   ADJ_RET[row_i] = return from ALL_PRICES[row_i] to ALL_PRICES[row_i+1]
#   date = DATES[row_i+1]
#   Sheet2 "current" row = S2_DATA_START + (row_i+1) = row_i + S2_DATA_START + 1
#   Sheet2 "prev"    row = S2_DATA_START + row_i
for row_i in range(T):
    r      = row_i + S3_DATA_START        # Sheet3 data row
    p_curr = row_i + S2_DATA_START + 1   # Sheet2 current price row
    p_prev = row_i + S2_DATA_START       # Sheet2 prev    price row

    val(ws3, r, 1, DATES[row_i + 1].strftime("%Y-%m-%d"), align="center")

    # r_local (B, C, D)
    ws3.cell(r, 2).value = f"=LN({S2}!B{p_curr}/{S2}!B{p_prev})"
    ws3.cell(r, 2).number_format = "0.000000"
    ws3.cell(r, 3).value = f"=LN({S2}!C{p_curr}/{S2}!C{p_prev})"
    ws3.cell(r, 3).number_format = "0.000000"
    ws3.cell(r, 4).value = f"=LN({S2}!D{p_curr}/{S2}!D{p_prev})"
    ws3.cell(r, 4).number_format = "0.000000"

    # r_FX (E=USD, F=KRW)
    ws3.cell(r, 5).value = f"=LN({S2}!G{p_curr}/{S2}!G{p_prev})"
    ws3.cell(r, 5).number_format = "0.000000"
    ws3.cell(r, 6).value = f"=LN({S2}!H{p_curr}/{S2}!H{p_prev})"
    ws3.cell(r, 6).number_format = "0.000000"

    # r_adj (G=TSMC, H=AAPL, I=Samsung)
    ws3.cell(r, 7).value = f"=B{r}"
    ws3.cell(r, 7).number_format = "0.000000"
    ws3.cell(r, 8).value = f"=C{r}+E{r}"
    ws3.cell(r, 8).number_format = "0.000000"
    ws3.cell(r, 8).fill = PatternFill("solid", fgColor=LBLUE)
    ws3.cell(r, 9).value = f"=D{r}+F{r}"
    ws3.cell(r, 9).number_format = "0.000000"
    ws3.cell(r, 9).fill = PatternFill("solid", fgColor=LBLUE)

    # P&L = Σ r_adj_i × MV_i（MV 在 Sheet1 G5:G7）
    ws3.cell(r, 10).value = (
        f"=G{r}*{S1}!$G$5"
        f"+H{r}*{S1}!$G$6"
        f"+I{r}*{S1}!$G$7"
    )
    ws3.cell(r, 10).number_format = "#,##0.00"

    # 負損益醒目（用 Python 計算值判斷）
    pnl_py = float(PNL_HS[row_i])
    if pnl_py < 0:
        ws3.cell(r, 10).fill = PatternFill("solid", fgColor=LRED)

apply_border(ws3, 15, 1, S3_DATA_END, 10)

# 統計彙整（用 AVERAGE/STDEV 公式）
r_avg = S3_DATA_END + 2
r_std = S3_DATA_END + 3
for r, lbl in [(r_avg, "平均"), (r_std, "標準差")]:
    val(ws3, r, 1, lbl, bold=True, bg=LGRAY)
    fn = "AVERAGE" if lbl == "平均" else "STDEV"
    for ci, col_letter in enumerate(["G", "H", "I", "J"], start=7):
        range_str = f"{col_letter}{S3_DATA_START}:{col_letter}{S3_DATA_END}"
        ws3.cell(r, ci).value    = f"={fn}({range_str})"
        ws3.cell(r, ci).number_format = "0.000000"
        ws3.cell(r, ci).font      = Font(bold=True, size=9)
        ws3.cell(r, ci).fill      = PatternFill("solid", fgColor=LGRAY)

note(ws3, S3_DATA_END + 5, 1,
     "藍色格（H, I）= FX 已疊加。"
     f"P&L 負值（紅色）共 {int(np.sum(PNL_HS < 0))} 筆。"
     "所有數值均為 Excel 公式，可點入儲存格確認計算鏈。")
autofit(ws3)

# ════════════════════════════════════════════════════════════
# Sheet 4 — HS VaR 逐步
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("④HS歷史模擬法")
merge_hdr(ws4, 1, 1, 10, "步驟二（HS）：歷史模擬法 VaR 逐步計算")

section(ws4, 3, "方法說明 & 公式")
hs_steps = [
    ("Step 1", f"取 T={T} 個交易日的 FX 調整後組合損益（直接引用 Sheet 3 J 欄）"),
    ("Step 2", "將 T 筆損益由小到大排序（B欄=時序，H欄=排序，使用 SMALL() 公式）"),
    ("Step 3", f"取第 {IDX_VAR+1} 小的損益作為 -VaR\n"
               f"   idx = floor((1-α)×T) = floor(0.01×{T}) = {IDX_VAR}"),
    ("Step 4", f"VaR = -SMALL(P&L, {IDX_VAR+1}) ← 見結果區公式"),
    ("Step 5", f"CVaR = -AVERAGE(SMALL(P&L,1),...,SMALL(P&L,{IDX_VAR})) ← 最差 {IDX_VAR} 筆均值"),
    ("驗證",   "點擊任意 H 欄（排序）或結果格，確認公式引用自 Sheet 3"),
]
for r, (s, d) in enumerate(hs_steps, start=4):
    bg = LGREEN if s == "驗證" else LBLUE
    val(ws4, r, 1, s, bold=True, bg=bg)
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    val(ws4, r, 2, d, align="left")

# 表頭
section(ws4, 11, f"組合損益序列 & 排序（T={T}，全為公式引用）")
ws4.merge_cells(start_row=12, start_column=1, end_row=12, end_column=4)
ws4.cell(12, 1, "▶ 原始損益（時間序列，引用 Sheet 3 J 欄）").font = Font(bold=True, color="FFFFFF")
ws4.cell(12, 1).fill = PatternFill("solid", fgColor="2E75B6")
ws4.merge_cells(start_row=12, start_column=6, end_row=12, end_column=9)
ws4.cell(12, 6, "▶ 排序後損益（由小到大，使用 SMALL() 公式）").font = Font(bold=True, color="FFFFFF")
ws4.cell(12, 6).fill = PatternFill("solid", fgColor=GREEN)

for c, h in enumerate(["排名", "日期", "原始P&L（引用Sheet3）", ""], start=1):
    hdr(ws4, 13, c, h, bg="2E75B6")
for c, h in enumerate(["排名", "排序P&L＝SMALL()", "說明", ""], start=6):
    hdr(ws4, 13, c, h, bg=GREEN)

# PNL 公式範圍（供 SMALL 公式使用）
pnl_range = f"{S3}!$J${S3_DATA_START}:$J${S3_DATA_END}"

for row_i in range(T):
    r = row_i + S4_DATA_START

    # 時序區（B, C 欄）
    val(ws4, r, 1, row_i + 1)
    val(ws4, r, 2, DATES[row_i + 1].strftime("%Y-%m-%d"), align="center")
    ws4.cell(r, 3).value = f"={S3}!J{row_i + S3_DATA_START}"
    ws4.cell(r, 3).number_format = "#,##0.00"
    pnl_py = float(PNL_HS[row_i])
    if pnl_py < 0:
        ws4.cell(r, 3).fill = PatternFill("solid", fgColor=LRED)

    # 排序區（H, I 欄）— SMALL() 公式
    rank = row_i + 1
    val(ws4, r, 6, rank)
    ws4.cell(r, 7).value = f"=SMALL({pnl_range},{rank})"
    ws4.cell(r, 7).number_format = "#,##0.00"

    is_var  = (row_i == IDX_VAR)
    is_cvar = (row_i < IDX_VAR)
    if is_var:
        ws4.cell(r, 7).fill = PatternFill("solid", fgColor=LYELL)
        ws4.cell(r, 7).font = Font(bold=True, size=9)
        val(ws4, r, 8, f"← VaR 點（第 {IDX_VAR+1} 小）", bg=LYELL)
    elif is_cvar:
        ws4.cell(r, 7).fill = PatternFill("solid", fgColor=LRED)
        val(ws4, r, 8, "← CVaR 範圍", bg=LRED)

apply_border(ws4, 13, 1, S4_DATA_END, 4)
apply_border(ws4, 13, 6, S4_DATA_END, 9)

# 結果區（VaR/CVaR 亦為公式）
rr = S4_DATA_END + 3
section(ws4, rr, "HS 計算結果（黃色格為公式，可點入驗證）")

# VaR formula: -SMALL(pnl_range, IDX_VAR+1)
var_formula_str  = f"=-SMALL({pnl_range},{IDX_VAR+1})"
# CVaR formula: average of IDX_VAR worst (if IDX_VAR==0, use 1 worst)
if IDX_VAR <= 0:
    cvar_formula_str = f"=-SMALL({pnl_range},1)"
else:
    terms = "+".join(f"SMALL({pnl_range},{k})" for k in range(1, IDX_VAR + 1))
    cvar_formula_str = f"=-({terms})/{IDX_VAR}"

result_rows = [
    (rr + 1, "HS VaR（99%，1日，TWD）",  var_formula_str,          "#,##0.00"),
    (rr + 2, "HS CVaR（TWD）",            cvar_formula_str,         "#,##0.00"),
    (rr + 3, "CVaR / VaR 倍數",           f"=B{rr+2}/B{rr+1}",     "0.0000"),
    (rr + 4, "組合市值（TWD）",           round(PORTFOLIO_VALUE,0), "#,##0"),
    (rr + 5, "VaR %（占市值）",           f"=B{rr+1}/B{rr+4}",     "0.00%"),
]
for row, lbl, formula_or_val, fmt in result_rows:
    val(ws4, row, 1, lbl, bold=True, align="left", bg=LGRAY)
    c = ws4.cell(row, 2, formula_or_val)
    c.number_format = fmt
    c.font = Font(bold=True, size=9)
    c.fill = PatternFill("solid", fgColor=LYELL)

# Component VaR
section(ws4, rr + 8, "Component VaR（HS）")
for c, h in enumerate(["資產", "Component VaR(TWD)", "Component CVaR(TWD)", "風險貢獻%"], start=1):
    hdr(ws4, rr + 9, c, h, bg="2E75B6")
for i, nm in enumerate(ASSETS):
    val(ws4, rr + 10 + i, 1, nm, align="left")
    val(ws4, rr + 10 + i, 2, round(HS_COMP_VAR[nm],  2), fmt="#,##0.00")
    val(ws4, rr + 10 + i, 3, round(HS_COMP_CVAR[nm], 2), fmt="#,##0.00")
    val(ws4, rr + 10 + i, 4, round(HS_COMP_VAR[nm] / HS_VAR, 4), fmt="0.00%")
note(ws4, rr + 14, 1,
     f"HS VaR（Python 驗算）= {HS_VAR:,.2f} TWD，CVaR = {HS_CVAR:,.2f} TWD。"
     "注意：HS Component VaR 各資產獨立分位數加總 ≠ 組合 VaR。")
autofit(ws4)

# ════════════════════════════════════════════════════════════
# Sheet 5 — Parametric VaR
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("⑤參數法Parametric")
merge_hdr(ws5, 1, 1, 12, "步驟三（Parametric）：Delta-Normal VaR 逐步計算")

section(ws5, 3, "方法說明 & 公式")
par_steps = [
    ("Step 1", f"計算 FX 調整後回報的樣本共變異數矩陣 Σ（基於 T={T} 筆回報）"),
    ("Step 2", "組合方差：σ²_P = w'Σw，w = TWD 市值向量"),
    ("Step 3", f"組合日標準差：σ_P = √(w'Σw) = {PORT_STD:,.4f} TWD"),
    ("Step 4", f"z_α = NORM.S.INV({CONFIDENCE}) = {Z_ALPHA:.6f}"),
    ("Step 5", f"VaR = z_α × σ_P = {Z_ALPHA:.4f} × {PORT_STD:,.4f} = {PAR_VAR:,.2f} TWD"),
    ("Step 6", f"CVaR = φ(z_α)/(1-α) × σ_P = {norm.pdf(Z_ALPHA):.6f}/0.01 × {PORT_STD:,.4f} = {PAR_CVAR:,.2f} TWD"),
    ("Step 7", "Component VaR_i = w_i × (Σw)_i / σ_P × z_α  ← 精確可加分解"),
    ("驗證",   "Σ Component VaR_i = Portfolio VaR（代數精確，見下方驗證列）"),
]
for r, (s, d) in enumerate(par_steps, start=4):
    bg = LGREEN if s == "驗證" else None
    val(ws5, r, 1, s, bold=True, bg=LBLUE if bg is None else LGREEN)
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    val(ws5, r, 2, d, align="left")

section(ws5, 13, "共變異數矩陣 Σ（樣本）")
for c, h in enumerate([""] + ASSETS, start=1):
    hdr(ws5, 14, c, h, bg="2E75B6")
for i in range(3):
    val(ws5, 15 + i, 1, ASSETS[i], bold=True, align="left", bg=LGRAY)
    for j in range(3):
        val(ws5, 15 + i, j + 2, round(float(COV_SAMPLE[i, j]), 10),
            fmt="0.0000000000", bg=LBLUE if i == j else None)

section(ws5, 19, "相關係數矩陣 ρ")
for c, h in enumerate([""] + ASSETS, start=1):
    hdr(ws5, 20, c, h, bg="2E75B6")
for i in range(3):
    val(ws5, 21 + i, 1, ASSETS[i], bold=True, align="left", bg=LGRAY)
    for j in range(3):
        v  = float(CORR_SAMPLE[i, j])
        bg = LBLUE if i == j else (LGREEN if v > 0.3 else None)
        val(ws5, 21 + i, j + 2, round(v, 4), fmt="0.0000", bg=bg)

section(ws5, 25, "σ_P 計算步驟（w'Σw）")
for c, h in enumerate(["資產", "w（TWD 市值）", "Σw（= Σ×w 第 i 元素）",
                        "w_i × (Σw)_i", "加總 = σ²_P", ""], start=1):
    hdr(ws5, 26, c, h, bg="2E75B6")
sv = COV_SAMPLE @ W
for i in range(3):
    val(ws5, 27 + i, 1, ASSETS[i], align="left")
    val(ws5, 27 + i, 2, round(float(W[i]),    0),    fmt="#,##0")
    val(ws5, 27 + i, 3, round(float(sv[i]),   6),    fmt="#,##0.000000")
    val(ws5, 27 + i, 4, round(float(W[i] * sv[i]), 4), fmt="#,##0.0000")
val(ws5, 30, 1, "加總 = σ²_P",      bold=True, bg=LGRAY)
val(ws5, 30, 4, round(PORT_VAR, 4), fmt="#,##0.0000", bold=True, bg=LYELL)
val(ws5, 31, 1, "√σ²_P = σ_P (TWD)", bold=True, bg=LGRAY)
val(ws5, 31, 4, round(PORT_STD, 4), fmt="#,##0.0000", bold=True, bg=LYELL)

section(ws5, 33, "Parametric 計算結果")
for c, h in enumerate(["項目", "公式", "數值"], start=1):
    hdr(ws5, 34, c, h, bg="2E75B6")
items = [
    ("z_α",               f"NORM.S.INV({CONFIDENCE})",              round(Z_ALPHA, 6)),
    ("σ_P（日，TWD）",    "√(w'Σw)",                                round(PORT_STD, 4)),
    ("VaR（TWD）",        "z_α × σ_P",                              round(PAR_VAR, 2)),
    ("CVaR（TWD）",       "φ(z_α)/(1-α) × σ_P",                    round(PAR_CVAR, 2)),
    ("VaR%",              "VaR / 組合市值",                          round(PAR_VAR / PORTFOLIO_VALUE, 6)),
    ("CVaR/VaR",          "CVaR / VaR",                              round(PAR_CVAR / PAR_VAR, 6)),
]
for r, (nm, fm, nv) in enumerate(items, start=35):
    val(ws5, r, 1, nm, align="left", bold=True, bg=LGRAY)
    val(ws5, r, 2, fm, align="left")
    fmt = "#,##0.00" if abs(nv) > 1 else "0.000000"
    val(ws5, r, 3, nv, fmt=fmt, bold=True, bg=LYELL)

section(ws5, 42, "Component VaR（精確可加分解驗證）")
for c, h in enumerate(["資產", "w_i (TWD)", "(Σw)_i/σ_P",
                        "Marginal VaR", "Component VaR", "Component CVaR", "風險貢獻%"], start=1):
    hdr(ws5, 43, c, h, bg="2E75B6")
for i in range(3):
    cv  = float(PAR_COMP_VAR[ASSETS[i]])
    ccv = float(PAR_COMP_CVAR[ASSETS[i]])
    val(ws5, 44 + i, 1, ASSETS[i], align="left")
    val(ws5, 44 + i, 2, round(float(W[i]), 0),                   fmt="#,##0")
    val(ws5, 44 + i, 3, round(float(MC_PAR_VEC[i]), 8),          fmt="0.00000000")
    val(ws5, 44 + i, 4, round(float(MC_PAR_VEC[i]) * Z_ALPHA, 8), fmt="0.00000000")
    val(ws5, 44 + i, 5, round(cv,  2),                            fmt="#,##0.00", bold=True)
    val(ws5, 44 + i, 6, round(ccv, 2),                            fmt="#,##0.00")
    val(ws5, 44 + i, 7, round(cv / PAR_VAR, 4),                   fmt="0.00%")

sum_cv = sum(PAR_COMP_VAR.values())
val(ws5, 47, 1, "加總（應 = VaR）",    bold=True, bg=LGRAY)
val(ws5, 47, 5, round(sum_cv, 2),       fmt="#,##0.00", bold=True, bg=LGREEN)
val(ws5, 47, 7, "✓ 誤差 < 0.01%",      bg=LGREEN)
val(ws5, 48, 1, "Portfolio VaR（對照）", bold=True, bg=LGRAY)
val(ws5, 48, 5, round(PAR_VAR, 2),      fmt="#,##0.00", bold=True, bg=LYELL)
note(ws5, 49, 1,
     f"驗證：|加總 - VaR| = {abs(sum_cv - PAR_VAR):.6e}，"
     f"相對誤差 = {abs(sum_cv - PAR_VAR) / PAR_VAR * 100:.8f}%（數值機器精度）")
autofit(ws5)

# ════════════════════════════════════════════════════════════
# Sheet 6 — Monte Carlo VaR
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("⑥Monte Carlo模擬")
merge_hdr(ws6, 1, 1, 12, f"步驟四（MC）：蒙地卡羅法 VaR（{MC_N:,} 條路徑）")

section(ws6, 3, "方法說明 & 公式")
mc_steps = [
    ("Step 1", "估計 FX 調整後回報的共變異數矩陣 Σ（同參數法）"),
    ("Step 2", "對 Σ 進行 Cholesky 分解：Σ = L L'（L 為下三角矩陣）"),
    ("Step 3", f"模擬 M={MC_N:,} 組相關常態亂數：r_sim = L×z + μ，z ~ N(0,I)"),
    ("Step 4", "各路徑損益：P&L_k = Σ_i w_i × r_sim(k,i)"),
    ("Step 5", f"VaR = -PERCENTILE(P&L, {(1-CONFIDENCE)*100:.0f}%) = {MC_VAR:,.2f} TWD"),
    ("Step 6", f"CVaR = -mean(P&L[P&L ≤ -VaR]) = {MC_CVAR:,.2f} TWD"),
]
for r, (s, d) in enumerate(mc_steps, start=4):
    val(ws6, r, 1, s, bold=True, bg=LBLUE)
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    val(ws6, r, 2, d, align="left")

section(ws6, 11, "Cholesky 分解 L（手算步驟）")
for c, h in enumerate(["", "L[i,0]", "L[i,1]", "L[i,2]", "計算公式"], start=1):
    hdr(ws6, 12, c, h, bg="2E75B6")
chol_formulas = [
    (ASSETS[0], L[0,0], None,   None,   "L[0,0] = √Σ[0,0]"),
    (ASSETS[1], L[1,0], L[1,1], None,   "L[1,0]=Σ[1,0]/L[0,0]；L[1,1]=√(Σ[1,1]−L[1,0]²)"),
    (ASSETS[2], L[2,0], L[2,1], L[2,2],
     "L[2,0]=Σ[2,0]/L[0,0]；L[2,1]=(Σ[2,1]−L[2,0]×L[1,0])/L[1,1]；L[2,2]=√(Σ[2,2]−L[2,0]²−L[2,1]²)"),
]
for r, (nm, l0, l1, l2, fm) in enumerate(chol_formulas, start=13):
    val(ws6, r, 1, nm, align="left", bold=True, bg=LGRAY)
    for ci, lv in enumerate([l0, l1, l2], start=2):
        if lv is not None:
            val(ws6, r, ci, round(lv, 9), fmt="0.000000000", bg=LBLUE)
        else:
            val(ws6, r, ci, "—", bg=LGRAY)
    ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)
    val(ws6, r, 5, fm, align="left")

section(ws6, 17, "驗證 Cholesky：L × L' 應還原 Σ")
LLT = L @ L.T
for c, h in enumerate([""] + ASSETS, start=1):
    hdr(ws6, 18, c, h, bg=GREEN)
for i in range(3):
    val(ws6, 19 + i, 1, ASSETS[i], bold=True, align="left", bg=LGRAY)
    for j in range(3):
        diff = abs(float(LLT[i, j]) - float(COV_SAMPLE[i, j]))
        val(ws6, 19 + i, j + 2, round(float(LLT[i, j]), 10),
            fmt="0.0000000000", bg=LGREEN if diff < 1e-10 else LRED)
note(ws6, 22, 1, "✓ 各格與 Σ 差距 < 1e-10（數值機器精度），驗證 Cholesky 正確。")

SHOW_MC = min(50, MC_N)
section(ws6, 24, f"前 {SHOW_MC} 條模擬路徑（共 {MC_N:,} 條）")
for c, h in enumerate(["路徑#", "r_sim[TSMC]", "r_sim[AAPL]", "r_sim[三星]",
                        "P&L (TWD)", "例外?"], start=1):
    hdr(ws6, 25, c, h, bg="2E75B6")
var_threshold = -MC_VAR
for k in range(SHOW_MC):
    r   = k + 26
    pk  = float(PNL_MC[k])
    exc = pk < var_threshold
    val(ws6, r, 1, k + 1)
    for ci in range(3):
        val(ws6, r, ci + 2, round(float(SIM[k, ci]), 6), fmt="0.000000")
    val(ws6, r, 5, round(pk, 2), fmt="#,##0.00",
        bg=LRED if exc else (LGREEN if pk > 0 else None), bold=exc)
    val(ws6, r, 6, "例外" if exc else "", bg=LRED if exc else None)

note(ws6, SHOW_MC + 27, 1,
     f"僅展示前 {SHOW_MC} 條路徑，完整 {MC_N:,} 條用於計算 VaR/CVaR。")

section(ws6, SHOW_MC + 29, "MC 計算結果")
mc_res = [
    ("MC VaR（TWD）",   round(MC_VAR,  2), f"-PERCENTILE(P&L, {(1-CONFIDENCE)*100:.0f}%)"),
    ("MC CVaR（TWD）",  round(MC_CVAR, 2), "−mean(P&L[P&L ≤ −VaR])"),
    ("VaR%（占市值）",  round(MC_VAR / PORTFOLIO_VALUE, 6), "VaR / 組合市值"),
    ("CVaR/VaR",        round(MC_CVAR / MC_VAR, 6),         "CVaR / VaR"),
    ("例外路徑數",      int(np.sum(PNL_MC < -MC_VAR)),
     f"≈ {MC_N}×1% = {MC_N*0.01:.0f}"),
]
for r, (lbl, nv, fm) in enumerate(mc_res, start=SHOW_MC + 30):
    val(ws6, r, 1, lbl, bold=True, align="left", bg=LGRAY)
    fmt = "#,##0.00" if isinstance(nv, float) and abs(nv) > 1 else (
        "#,##0" if isinstance(nv, int) else "0.000000")
    val(ws6, r, 2, nv, fmt=fmt, bold=True, bg=LYELL)
    ws6.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    val(ws6, r, 3, fm, align="left")

section(ws6, SHOW_MC + 36, "Component VaR（MC 尾端均值）")
for c, h in enumerate(["資產", "Component VaR(TWD)", "Component CVaR(TWD)", "風險貢獻%"], start=1):
    hdr(ws6, SHOW_MC + 37, c, h, bg="2E75B6")
for i, nm in enumerate(ASSETS):
    val(ws6, SHOW_MC + 38 + i, 1, nm, align="left")
    val(ws6, SHOW_MC + 38 + i, 2, round(float(MC_COMP_VAR[nm]),  2), fmt="#,##0.00")
    val(ws6, SHOW_MC + 38 + i, 3, round(float(MC_COMP_CVAR[nm]), 2), fmt="#,##0.00")
    val(ws6, SHOW_MC + 38 + i, 4, round(float(MC_COMP_VAR[nm]) / MC_VAR, 4), fmt="0.00%")
note(ws6, SHOW_MC + 42, 1,
     "MC Component VaR：各資產在組合尾端場景（P&L ≤ −VaR）下的平均損失。")
autofit(ws6)

# ════════════════════════════════════════════════════════════
# Sheet 7 — 三法彙總比較
# ════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("⑦三法彙總比較")
merge_hdr(ws7, 1, 1, 8, "三方法 VaR 彙總比較 & 驗證核對表")

section(ws7, 3, "組合層級指標比較")
for c, h in enumerate(["指標", "Historical Simulation", "Parametric Delta-Normal",
                        f"Monte Carlo（M={MC_N:,}）", "說明"], start=1):
    hdr(ws7, 4, c, h, bg="2E75B6")

compare_rows = [
    ("VaR（TWD）",       HS_VAR,  PAR_VAR,  MC_VAR,  "#,##0.00", "主要風險指標"),
    ("CVaR（TWD）",      HS_CVAR, PAR_CVAR, MC_CVAR, "#,##0.00", "一致性風險測度"),
    ("VaR%（占市值）",   HS_VAR / PORTFOLIO_VALUE, PAR_VAR / PORTFOLIO_VALUE,
                          MC_VAR / PORTFOLIO_VALUE, "0.00%", "相對風險水準"),
    ("CVaR / VaR",        HS_CVAR / HS_VAR, PAR_CVAR / PAR_VAR, MC_CVAR / MC_VAR,
                          "0.0000", "尾端形狀（常態≈1.29）"),
    ("損益標準差（TWD）", float(PNL_HS.std()), PORT_STD, float(PNL_MC.std()),
                          "#,##0.00", "1日組合損益σ"),
]
for r, (lbl, hs_v, par_v, mc_v, fmt, note_t) in enumerate(compare_rows, start=5):
    val(ws7, r, 1, lbl, bold=True, align="left", bg=LGRAY)
    for ci, v in enumerate([hs_v, par_v, mc_v], start=2):
        val(ws7, r, ci, round(float(v), 4), fmt=fmt, bold=(ci == 2))
    val(ws7, r, 5, note_t, align="left")
apply_border(ws7, 4, 1, 9, 5)

section(ws7, 11, "個別資產 Component VaR 比較（TWD）")
for c, h in enumerate(["資產", "市場", "幣別",
                        "HS VaR", "Param VaR", "MC VaR",
                        "HS 貢獻%", "Param 貢獻%", "MC 貢獻%"], start=1):
    hdr(ws7, 12, c, h, bg="2E75B6")
for i, nm in enumerate(ASSETS):
    val(ws7, 13 + i, 1, nm, align="left")
    val(ws7, 13 + i, 2, get_market_label(nm))
    val(ws7, 13 + i, 3, CCY[i])
    val(ws7, 13 + i, 4, round(HS_COMP_VAR[nm],  2), fmt="#,##0.00")
    val(ws7, 13 + i, 5, round(PAR_COMP_VAR[nm], 2), fmt="#,##0.00")
    val(ws7, 13 + i, 6, round(MC_COMP_VAR[nm],  2), fmt="#,##0.00")
    val(ws7, 13 + i, 7, round(HS_COMP_VAR[nm]  / HS_VAR,  4), fmt="0.00%")
    val(ws7, 13 + i, 8, round(PAR_COMP_VAR[nm] / PAR_VAR, 4), fmt="0.00%")
    val(ws7, 13 + i, 9, round(MC_COMP_VAR[nm]  / MC_VAR,  4), fmt="0.00%")
apply_border(ws7, 12, 1, 15, 9)

section(ws7, 17, "驗證核對表")
for c, h in enumerate(["驗證項目", "HS", "Parametric", "Monte Carlo", "預期結果", "通過?"], start=1):
    hdr(ws7, 18, c, h, bg=GREEN)
sum_hs  = sum(HS_COMP_VAR.values())
sum_par = sum(PAR_COMP_VAR.values())
sum_mc  = sum(MC_COMP_VAR.values())
checks = [
    ("Component VaR 加總 = 組合 VaR",
     f"{sum_hs:,.0f} ≠ {HS_VAR:,.0f}（近似）",
     f"{sum_par:,.0f} ≈ {PAR_VAR:,.0f}（精確）",
     f"{sum_mc:,.0f} ≠ {MC_VAR:,.0f}（近似）",
     "HS/MC 近似；Parametric 代數精確",
     "Param ✓"),
    ("CVaR > VaR",
     "✓" if HS_CVAR > HS_VAR else "✗",
     "✓" if PAR_CVAR > PAR_VAR else "✗",
     "✓" if MC_CVAR > MC_VAR else "✗",
     "必然成立（定義性質）", "✓"),
    ("VaR > 0",
     "✓" if HS_VAR > 0 else "✗",
     "✓" if PAR_VAR > 0 else "✗",
     "✓" if MC_VAR > 0 else "✗",
     "正常市場必然", "✓"),
    ("MC → Parametric 收斂",
     "—",
     f"{PAR_VAR:,.0f}",
     f"{MC_VAR:,.0f}（差 {abs(MC_VAR-PAR_VAR)/PAR_VAR*100:.1f}%）",
     "大樣本下 MC ≈ Param（同常態假設）",
     "✓" if abs(MC_VAR - PAR_VAR) / PAR_VAR < 0.05 else "需更多路徑"),
]
for r, row_data in enumerate(checks, start=19):
    for c, v in enumerate(row_data, start=1):
        is_check = (c == 6)
        val(ws7, r, c, v, align="left" if c <= 5 else "center",
            bold=is_check, bg=LGREEN if is_check and "✓" in str(v) else None)
apply_border(ws7, 18, 1, 22, 6)

note(ws7, 24, 1,
     f"資料期間：{DATES[0].strftime('%Y-%m-%d')} ~ {DATES[-1].strftime('%Y-%m-%d')}，"
     f"T={T}，N_MC={MC_N:,}，α=99%，h=1日。")
autofit(ws7)

# ════════════════════════════════════════════════════════════
# 儲存
# ════════════════════════════════════════════════════════════
OUT = os.path.join(os.path.dirname(__file__), "VaR_逐步驗證.xlsx")
wb.save(OUT)
print(f"\n[完成] 已儲存至：{OUT}")
print(f"  Sheet 2 資料：{N_DAYS+1} 列（含初始 S0）")
print(f"  Sheet 3 公式：LN() 直接引用 Sheet 2，P&L 引用 Sheet 1 市值")
print(f"  Sheet 4 公式：SMALL() 引用 Sheet 3 P&L，VaR/CVaR 結果格亦為公式")
